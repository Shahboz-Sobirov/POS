# -*- coding: utf-8 -*-
"""
Reports Page (F4)
Daily, Weekly, Monthly, Yearly, Custom Range
"""
from html import escape
from datetime import datetime, timedelta
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QComboBox,
    QDateEdit,
    QGroupBox,
    QFileDialog,
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QShortcut, QKeySequence, QTextDocument
from PySide6.QtPrintSupport import QPrinter, QPrintDialog, QPrinterInfo
from openpyxl import Workbook
from openpyxl.styles import Font
from config.constants import APP_NAME
from services.sale_service import SaleService
from services.debt_payment_service import DebtPaymentService
from utils.formatter import format_meters, format_square_meters
from widgets.custom_alert import CustomAlert
from utils.error_logger import log_exception, get_user_friendly_message


class ReportsPage(QWidget):
    def __init__(self):
        super().__init__()
        self.current_sales = []
        self.current_debt_payments = []
        self.current_records = []
        self.current_start_datetime = None
        self.current_end_datetime = None
        self.current_payment_totals = self.create_empty_payment_totals()
        self.current_total_kvm = 0
        self.current_top_glass = "-"
        self.setup_ui()
        self.setup_shortcuts()

    @staticmethod
    def create_empty_payment_totals():
        """Create empty payment totals structure."""
        return {
            "naqd": 0,
            "karta": 0,
            "click": 0,
            "qarz": 0,
            "tolangan_qarz": 0,
        }

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(32, 32, 32, 32)
        layout.setSpacing(16)

        title = QLabel("Hisobotlar")
        title.setObjectName("pageTitle")
        layout.addWidget(title)

        filter_group = QGroupBox("Filtr")
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(12)

        mode_label = QLabel("Davr:")
        filter_layout.addWidget(mode_label)

        self.mode_combo = QComboBox()
        self.mode_combo.setMinimumHeight(40)
        self.mode_combo.addItem("Kunlik", "daily")
        self.mode_combo.addItem("Haftalik", "weekly")
        self.mode_combo.addItem("Oylik", "monthly")
        self.mode_combo.addItem("Yillik", "yearly")
        self.mode_combo.addItem("Maxsus", "custom")
        self.mode_combo.currentIndexChanged.connect(self.on_mode_changed)
        filter_layout.addWidget(self.mode_combo)

        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(QDate.currentDate())
        self.start_date.setMinimumHeight(40)
        self.start_date.setDisplayFormat("dd.MM.yyyy")
        self.apply_date_style(self.start_date)
        filter_layout.addWidget(self.start_date)

        range_label = QLabel("-")
        filter_layout.addWidget(range_label)

        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setMinimumHeight(40)
        self.end_date.setDisplayFormat("dd.MM.yyyy")
        self.end_date.setEnabled(False)
        self.apply_date_style(self.end_date)
        filter_layout.addWidget(self.end_date)

        load_btn = QPushButton("Yuklash")
        load_btn.setObjectName("btnSuccess")
        load_btn.setMinimumHeight(40)
        load_btn.setCursor(Qt.PointingHandCursor)
        load_btn.clicked.connect(self.load_report)
        filter_layout.addWidget(load_btn)

        print_btn = QPushButton("Chop Etish")
        print_btn.setMinimumHeight(40)
        print_btn.setCursor(Qt.PointingHandCursor)
        print_btn.clicked.connect(self.print_report)
        filter_layout.addWidget(print_btn)

        excel_btn = QPushButton("Excel")
        excel_btn.setMinimumHeight(40)
        excel_btn.setCursor(Qt.PointingHandCursor)
        excel_btn.clicked.connect(self.export_excel)
        filter_layout.addWidget(excel_btn)

        filter_layout.addStretch()
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)

        summary_group = QGroupBox("Umumiy Ma'lumot")
        summary_layout = QHBoxLayout()
        summary_layout.setSpacing(16)

        self.total_sales_label = QLabel("Savdolar: 0 | Qarz to'lovlari: 0")
        self.total_sales_label.setStyleSheet("font-size: 16px; font-weight: 600;")
        summary_layout.addWidget(self.total_sales_label)

        self.total_revenue_label = QLabel("Daromad: 0 so'm")
        self.total_revenue_label.setStyleSheet(
            "font-size: 16px; font-weight: 600; color: #3498db;"
        )
        summary_layout.addWidget(self.total_revenue_label)

        self.total_profit_label = QLabel("Foyda: 0 so'm")
        self.total_profit_label.setStyleSheet(
            "font-size: 16px; font-weight: 600; color: #27ae60;"
        )
        summary_layout.addWidget(self.total_profit_label)

        self.total_kvm_label = QLabel("KVM: 0")
        self.total_kvm_label.setStyleSheet(
            "font-size: 16px; font-weight: 600; color: #0f766e;"
        )
        summary_layout.addWidget(self.total_kvm_label)

        self.top_glass_label = QLabel("Eng ko'p: -")
        self.top_glass_label.setStyleSheet(
            "font-size: 16px; font-weight: 600; color: #102331;"
        )
        summary_layout.addWidget(self.top_glass_label)

        summary_layout.addStretch()
        summary_group.setLayout(summary_layout)
        layout.addWidget(summary_group)

        payment_group = QGroupBox("To'lovlar Taqsimoti")
        payment_layout = QHBoxLayout()
        payment_layout.setSpacing(16)

        self.naqd_label = QLabel("Naqd: 0")
        payment_layout.addWidget(self.naqd_label)

        self.karta_label = QLabel("Karta: 0")
        payment_layout.addWidget(self.karta_label)

        self.click_label = QLabel("Click: 0")
        payment_layout.addWidget(self.click_label)

        self.qarz_label = QLabel("Qarz: 0")
        payment_layout.addWidget(self.qarz_label)

        self.tolangan_qarz_label = QLabel("To'langan qarz: 0")
        payment_layout.addWidget(self.tolangan_qarz_label)

        payment_layout.addStretch()
        payment_group.setLayout(payment_layout)
        layout.addWidget(payment_group)

        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(
            ["Sana", "Mijoz", "Oyna", "Eni", "Bo'yi", "KVM", "To'lov", "Jami", "Foyda"]
        )
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(48)
        layout.addWidget(self.table)

        self.setLayout(layout)
        self.load_report()

    def setup_shortcuts(self):
        """Setup shortcuts"""
        QShortcut(QKeySequence("Ctrl+P"), self).activated.connect(self.print_report)
        QShortcut(QKeySequence("Ctrl+E"), self).activated.connect(self.export_excel)

    def on_mode_changed(self):
        """Mode changed"""
        mode = self.mode_combo.currentData()
        self.end_date.setEnabled(mode == "custom")

        if mode == "custom":
            if self.end_date.date() < self.start_date.date():
                self.end_date.setDate(self.start_date.date())
            return

        today = datetime.now().date()
        if mode == "daily":
            start_date = today
            end_date = today
        elif mode == "weekly":
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif mode == "monthly":
            start_date = today.replace(day=1)
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
        else:
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)

        self.start_date.setDate(QDate(start_date.year, start_date.month, start_date.day))
        self.end_date.setDate(QDate(end_date.year, end_date.month, end_date.day))
        self.load_report()

    def get_date_range(self):
        """Resolve current date range based on selected mode."""
        mode = self.mode_combo.currentData()
        start_date = self.start_date.date().toPython()

        if mode == "custom":
            end_date = self.end_date.date().toPython()
            if end_date < start_date:
                raise ValueError("Tugash sanasi boshlanish sanasidan kichik bo'lishi mumkin emas")
        elif mode == "daily":
            end_date = start_date
        elif mode == "weekly":
            end_date = start_date + timedelta(days=6)
        elif mode == "monthly":
            next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            end_date = next_month - timedelta(days=1)
        else:
            end_date = start_date.replace(month=12, day=31)

        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        return start_datetime, end_datetime

    def load_report(self):
        """Load report."""
        try:
            start_datetime, end_datetime = self.get_date_range()

            sales = SaleService.get_by_date_range(start_datetime, end_datetime)
            debt_payments = DebtPaymentService.get_by_date_range(start_datetime, end_datetime)
            payment_totals = self.create_empty_payment_totals()

            for sale in sales:
                if not sale.payment_breakdown:
                    continue
                for key, value in sale.payment_breakdown.items():
                    if key in payment_totals:
                        payment_totals[key] += float(value or 0)

            payment_totals["tolangan_qarz"] = sum(
                payment.amount or 0 for payment in debt_payments
            )

            self.current_sales = sales
            self.current_debt_payments = debt_payments
            self.current_records = self.build_report_records(sales, debt_payments)
            self.current_start_datetime = start_datetime
            self.current_end_datetime = end_datetime
            self.current_payment_totals = payment_totals
            self.current_total_kvm = sum(record.get("kvm", 0) for record in self.current_records if record["row_type"] == "sale")
            self.current_top_glass = self.get_top_glass_name(self.current_records)

            total_sales = len(sales)
            total_debt_payments = len(debt_payments)
            total_revenue = sum(s.total_amount for s in sales)
            total_profit = sum(s.profit for s in sales)

            self.total_sales_label.setText(
                f"Savdolar: {total_sales} | Qarz to'lovlari: {total_debt_payments}"
            )
            self.total_revenue_label.setText(f"Daromad: {total_revenue:,.0f} so'm")
            self.total_profit_label.setText(f"Foyda: {total_profit:,.0f} so'm")
            self.total_kvm_label.setText(f"KVM: {format_square_meters(self.current_total_kvm)}")
            self.top_glass_label.setText(f"Eng ko'p: {self.current_top_glass}")

            self.naqd_label.setText(f"Naqd: {payment_totals['naqd']:,.0f}")
            self.karta_label.setText(f"Karta: {payment_totals['karta']:,.0f}")
            self.click_label.setText(f"Click: {payment_totals['click']:,.0f}")
            self.qarz_label.setText(f"Qarz: {payment_totals['qarz']:,.0f}")
            self.tolangan_qarz_label.setText(
                f"To'langan qarz: {payment_totals['tolangan_qarz']:,.0f}"
            )

            self.populate_table(self.current_records)

        except Exception as e:
            log_exception(e, "load_report")
            CustomAlert.show_error(self, "Xato", get_user_friendly_message(e))

    def build_report_records(self, sales, debt_payments):
        """Build one table dataset from sales and debt payments."""
        records = []

        for sale in sales:
            customer_name = sale.customer.full_name if sale.customer else "-"
            payment_text = self.format_payment_breakdown(sale.payment_breakdown)
            for item in sale.items:
                if not item.product:
                    continue
                eni = item.eni or item.width
                boyi = item.boyi or item.height
                kvm = item.kvm or item.area_sqm or item.quantity or 0
                narx_per_kvm = item.narx_per_kvm or item.price or 0
                records.append(
                    {
                        "row_type": "sale",
                        "date": sale.sale_date,
                        "customer_name": customer_name,
                        "product_name": item.product.name,
                        "eni": eni,
                        "boyi": boyi,
                        "kvm": kvm,
                        "narx_per_kvm": narx_per_kvm,
                        "payment_text": payment_text,
                        "amount": kvm * narx_per_kvm,
                        "profit": item.profit or 0,
                        "name": item.product.name,
                    }
                )

        for payment in debt_payments:
            customer_name = payment.customer.full_name if payment.customer else "-"
            description = "To'langan qarz"
            if payment.note:
                description = f"{description}: {payment.note}"

            records.append(
                {
                    "row_type": "debt_payment",
                    "date": payment.payment_date,
                    "customer_name": customer_name,
                    "product_name": description,
                    "eni": None,
                    "boyi": None,
                    "kvm": 0,
                    "narx_per_kvm": 0,
                    "payment_text": self.format_payment_breakdown(
                        payment.payment_breakdown,
                        payment.payment_type,
                    ),
                    "amount": payment.amount or 0,
                    "profit": 0,
                    "name": "",
                }
            )

        records.sort(key=lambda record: record["date"], reverse=True)
        return records

    def get_top_glass_name(self, records):
        """Return the glass type with the highest sold KVM in the report."""
        totals = {}
        for record in records:
            if record["row_type"] != "sale":
                continue
            name = record.get("name") or "-"
            totals[name] = totals.get(name, 0) + float(record.get("kvm") or 0)

        if not totals:
            return "-"

        name, kvm = max(totals.items(), key=lambda item: item[1])
        return f"{name} ({format_square_meters(kvm)})"

    def populate_table(self, records):
        """Populate table with safe rendering."""
        self.table.setRowCount(len(records))

        for row, record in enumerate(records):
            try:
                date_str = record["date"].strftime("%d.%m.%Y %H:%M") if record["date"] else "-"
                self.table.setItem(row, 0, QTableWidgetItem(date_str))
                self.table.setItem(row, 1, QTableWidgetItem(record["customer_name"]))
                self.table.setItem(row, 2, QTableWidgetItem(record["product_name"]))
                self.table.setItem(row, 3, QTableWidgetItem(format_meters(record["eni"]) if record["eni"] else "-"))
                self.table.setItem(row, 4, QTableWidgetItem(format_meters(record["boyi"]) if record["boyi"] else "-"))
                self.table.setItem(
                    row,
                    5,
                    QTableWidgetItem(
                        format_square_meters(record["kvm"]) if record["row_type"] == "sale" else "-"
                    ),
                )
                self.table.setItem(row, 6, QTableWidgetItem(record["payment_text"]))
                self.table.setItem(row, 7, QTableWidgetItem(f"{record['amount']:,.0f}"))
                self.table.setItem(row, 8, QTableWidgetItem(f"{record['profit']:,.0f}"))
                for col in range(3, 9):
                    item = self.table.item(row, col)
                    if item:
                        item.setTextAlignment(Qt.AlignCenter)
            except Exception as e:
                print(f"Warning: Failed to render report row {row}: {e}")
                for col in range(9):
                    self.table.setItem(row, col, QTableWidgetItem("-"))
        self.table.resizeRowsToContents()

    def format_payment_breakdown(self, breakdown, fallback_payment_type=None):
        """Format payment breakdown without showing 'Mixed' text."""
        if not breakdown:
            return fallback_payment_type or "Naqd"

        parts = []
        if breakdown.get("naqd", 0) > 0:
            parts.append(f"Naqd: {breakdown['naqd']:,.0f}")
        if breakdown.get("karta", 0) > 0:
            parts.append(f"Karta: {breakdown['karta']:,.0f}")
        if breakdown.get("click", 0) > 0:
            parts.append(f"Click: {breakdown['click']:,.0f}")
        if breakdown.get("qarz", 0) > 0:
            parts.append(f"Qarz: {breakdown['qarz']:,.0f}")

        return "\n".join(parts) if parts else (fallback_payment_type or "Naqd")

    def build_report_html(self):
        """Build printable HTML from the currently loaded report."""
        total_revenue = sum(s.total_amount for s in self.current_sales)
        total_profit = sum(s.profit for s in self.current_sales)
        date_range = (
            f"{self.current_start_datetime.strftime('%d.%m.%Y')} - "
            f"{self.current_end_datetime.strftime('%d.%m.%Y')}"
        )

        rows = []
        for record in self.current_records:
            rows.append(
                f"""
                <tr>
                    <td>{escape(record['date'].strftime("%d.%m.%Y %H:%M"))}</td>
                    <td>{escape(record['customer_name'])}</td>
                    <td>{escape(record['product_name'])}</td>
                    <td>{escape(format_meters(record['eni']) if record['eni'] else "-")}</td>
                    <td>{escape(format_meters(record['boyi']) if record['boyi'] else "-")}</td>
                    <td>{escape(format_square_meters(record['kvm']) if record['row_type'] == 'sale' else "-")}</td>
                    <td>{escape(record['payment_text']).replace(chr(10), "<br>")}</td>
                    <td style="text-align:right;">{record['amount']:,.0f}</td>
                    <td style="text-align:right;">{record['profit']:,.0f}</td>
                </tr>
                """
            )

        return f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; font-size: 12px; color: #111827; }}
                h1 {{ font-size: 20px; margin-bottom: 4px; }}
                p {{ margin: 4px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
                th, td {{ border: 1px solid #CBD5E1; padding: 8px; vertical-align: top; }}
                th {{ background: #E2E8F0; text-align: left; }}
                .summary {{ margin-top: 12px; }}
            </style>
        </head>
        <body>
            <h1>{escape(APP_NAME)} Hisoboti</h1>
            <p>Davr: {escape(date_range)}</p>
            <div class="summary">
                <p>Savdolar: {len(self.current_sales)} | Qarz to'lovlari: {len(self.current_debt_payments)}</p>
                <p>Daromad: {total_revenue:,.0f} so'm</p>
                <p>Foyda: {total_profit:,.0f} so'm</p>
                <p>Sotilgan KVM: {escape(format_square_meters(self.current_total_kvm))}</p>
                <p>Eng ko'p sotilgan oyna: {escape(self.current_top_glass)}</p>
                <p>Naqd: {self.current_payment_totals['naqd']:,.0f}</p>
                <p>Karta: {self.current_payment_totals['karta']:,.0f}</p>
                <p>Click: {self.current_payment_totals['click']:,.0f}</p>
                <p>Qarz: {self.current_payment_totals['qarz']:,.0f}</p>
                <p>To'langan qarz: {self.current_payment_totals['tolangan_qarz']:,.0f}</p>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Sana</th>
                        <th>Mijoz</th>
                        <th>Oyna</th>
                        <th>Eni</th>
                        <th>Bo'yi</th>
                        <th>KVM</th>
                        <th>To'lov</th>
                        <th>Jami</th>
                        <th>Foyda</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
        </body>
        </html>
        """

    def print_report(self):
        """Print report (CTRL+P)."""
        if not self.current_records:
            CustomAlert.show_warning(self, "Ogohlantirish", "Chop etish uchun hisobot bo'sh")
            return

        if not self.check_printer_available():
            CustomAlert.show_warning(self, "Ogohlantirish", "Printer topilmadi")
            return

        try:
            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self)
            if not dialog.exec():
                return

            document = QTextDocument()
            document.setHtml(self.build_report_html())
            document.print(printer)
            CustomAlert.show_success(self, "Muvaffaqiyat", "Hisobot printerga yuborildi")
        except Exception as e:
            log_exception(e, "print_report")
            CustomAlert.show_error(self, "Xato", get_user_friendly_message(e))

    def check_printer_available(self):
        """Check if printer is available."""
        try:
            default_printer = QPrinterInfo.defaultPrinter()
            return not default_printer.isNull()
        except Exception:
            return False

    def export_excel(self):
        """Export to Excel (CTRL+E)."""
        if not self.current_records:
            CustomAlert.show_warning(self, "Ogohlantirish", "Export qilish uchun hisobot bo'sh")
            return

        default_name = (
            f"hisobot_{self.current_start_datetime.strftime('%Y%m%d')}_"
            f"{self.current_end_datetime.strftime('%Y%m%d')}.xlsx"
        )
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Excel faylni saqlash",
            default_name,
            "Excel Files (*.xlsx)",
        )

        if not file_path:
            return

        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Hisobot"

            sheet["A1"] = f"{APP_NAME} Hisoboti"
            sheet["A1"].font = Font(bold=True, size=14)
            sheet["A2"] = "Davr"
            sheet["B2"] = (
                f"{self.current_start_datetime.strftime('%d.%m.%Y')} - "
                f"{self.current_end_datetime.strftime('%d.%m.%Y')}"
            )
            sheet["A3"] = "Savdolar"
            sheet["B3"] = len(self.current_sales)
            sheet["A4"] = "Qarz to'lovlari"
            sheet["B4"] = len(self.current_debt_payments)
            sheet["A5"] = "Daromad"
            sheet["B5"] = sum(s.total_amount for s in self.current_sales)
            sheet["A6"] = "Foyda"
            sheet["B6"] = sum(s.profit for s in self.current_sales)
            sheet["A7"] = "Sotilgan KVM"
            sheet["B7"] = self.current_total_kvm
            sheet["A8"] = "Eng ko'p sotilgan oyna"
            sheet["B8"] = self.current_top_glass
            sheet["A9"] = "Naqd"
            sheet["B9"] = self.current_payment_totals["naqd"]
            sheet["A10"] = "Karta"
            sheet["B10"] = self.current_payment_totals["karta"]
            sheet["A11"] = "Click"
            sheet["B11"] = self.current_payment_totals["click"]
            sheet["A12"] = "Qarz"
            sheet["B12"] = self.current_payment_totals["qarz"]
            sheet["A13"] = "To'langan qarz"
            sheet["B13"] = self.current_payment_totals["tolangan_qarz"]

            header_row = 15
            headers = ["Sana", "Mijoz", "Oyna", "Eni", "Bo'yi", "KVM", "To'lov", "Jami", "Foyda"]
            for column, header in enumerate(headers, start=1):
                cell = sheet.cell(row=header_row, column=column, value=header)
                cell.font = Font(bold=True)

            for row_index, record in enumerate(self.current_records, start=header_row + 1):
                sheet.cell(row=row_index, column=1, value=record["date"].strftime("%d.%m.%Y %H:%M"))
                sheet.cell(row=row_index, column=2, value=record["customer_name"])
                sheet.cell(row=row_index, column=3, value=record["product_name"])
                sheet.cell(row=row_index, column=4, value=format_meters(record["eni"]) if record["eni"] else "-")
                sheet.cell(row=row_index, column=5, value=format_meters(record["boyi"]) if record["boyi"] else "-")
                sheet.cell(
                    row=row_index,
                    column=6,
                    value=format_square_meters(record["kvm"]) if record["row_type"] == "sale" else "-",
                )
                sheet.cell(row=row_index, column=7, value=record["payment_text"])
                sheet.cell(row=row_index, column=8, value=record["amount"])
                sheet.cell(row=row_index, column=9, value=record["profit"])

            sheet.freeze_panes = "A16"
            sheet.column_dimensions["A"].width = 18
            sheet.column_dimensions["B"].width = 24
            sheet.column_dimensions["C"].width = 28
            sheet.column_dimensions["D"].width = 12
            sheet.column_dimensions["E"].width = 12
            sheet.column_dimensions["F"].width = 14
            sheet.column_dimensions["G"].width = 28
            sheet.column_dimensions["H"].width = 16
            sheet.column_dimensions["I"].width = 16

            workbook.save(file_path)
            CustomAlert.show_success(self, "Muvaffaqiyat", f"Excel fayl saqlandi:\n{file_path}")
        except Exception as e:
            log_exception(e, "export_excel")
            CustomAlert.show_error(self, "Xato", get_user_friendly_message(e))

    def refresh(self):
        """Refresh page."""
        self.load_report()

    def apply_date_style(self, date_edit):
        """Apply custom style to QDateEdit and QCalendarWidget."""
        date_edit.setStyleSheet(
            """
            QDateEdit {
                background-color: white;
                border: 1px solid #CBD5E1;
                border-radius: 8px;
                padding-left: 10px;
                color: #0F172A;
                font-size: 13px;
            }
            QDateEdit:hover {
                border: 1px solid #38BDF8;
            }
            QDateEdit:focus {
                border: 2px solid #38BDF8;
            }
            QDateEdit:disabled {
                background-color: #F1F5F9;
                color: #94A3B8;
            }
            QDateEdit::drop-down {
                border: none;
                width: 30px;
            }
            QDateEdit::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid #0F172A;
                margin-right: 8px;
            }
            QDateEdit:disabled::down-arrow {
                border-top-color: #94A3B8;
            }

            QCalendarWidget {
                background-color: #F8FAFC;
                border: 1px solid #CBD5E1;
                border-radius: 10px;
            }
            QCalendarWidget QToolButton {
                background-color: transparent;
                color: #0F172A;
                border-radius: 4px;
                padding: 4px;
                font-weight: 600;
            }
            QCalendarWidget QToolButton:hover {
                background-color: #E0F2FE;
            }
            QCalendarWidget QMenu {
                background-color: white;
                border: 1px solid #CBD5E1;
                color: #0F172A;
            }
            QCalendarWidget QSpinBox {
                background-color: white;
                border: 1px solid #CBD5E1;
                border-radius: 4px;
                padding: 4px;
                color: #0F172A;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #082F49;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
            }
            QCalendarWidget QWidget {
                color: #0F172A;
            }
            QCalendarWidget QAbstractItemView {
                background-color: white;
                selection-background-color: #38BDF8;
                selection-color: white;
                color: #0F172A;
            }
            QCalendarWidget QAbstractItemView:enabled {
                color: #0F172A;
            }
            QCalendarWidget QAbstractItemView:disabled {
                color: #94A3B8;
            }
            """
        )
